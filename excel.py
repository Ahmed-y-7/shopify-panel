"""Export shopify.db to a dashboard Excel workbook (shopify.xlsx).

Sheets: Dashboard (KPIs + charts, live formulas), Orders, Items, Inventory.
Everything aggregates via SUMIFS/COUNTIFS over the data sheets, so editing
the data (or the blue month cell) recalculates the dashboard in Excel.
"""
import datetime
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ARIAL = Font(name="Arial", size=11)
BOLD = Font(name="Arial", size=11, bold=True)
TITLE = Font(name="Arial", size=16, bold=True)
INPUT = Font(name="Arial", size=11, color="0000FF", bold=True)  # editable cells
NUM = "#,##0"


def _sheet(wb, name, headers, rows, widths, formats=None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.font = BOLD
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = ARIAL
            if formats and cell.column_letter in formats:
                cell.number_format = formats[cell.column_letter]
    ws.freeze_panes = "A2"
    return ws


def export(db):
    orders = db.execute("""SELECT name, created_at,
                           substr(datetime(created_at,'+3 hours'),1,10),
                           substr(datetime(created_at,'+3 hours'),1,7),
                           financial_status, total
                           FROM orders ORDER BY created_at""").fetchall()
    items = db.execute("""SELECT o.name, i.title, i.quantity, i.price
                          FROM order_items i JOIN orders o ON o.id=i.order_id
                          ORDER BY o.created_at""").fetchall()
    inventory = db.execute("""SELECT p.title, p.status, v.title, v.sku,
                              v.price, v.inventory
                              FROM variants v JOIN products p ON p.id=v.product_id
                              ORDER BY p.title, v.title""").fetchall()
    months = [r[0] for r in db.execute(
        "SELECT DISTINCT substr(datetime(created_at,'+3 hours'),1,7) m "
        "FROM orders ORDER BY m")]
    days = db.execute(
        "SELECT substr(datetime(created_at,'+3 hours'),1,10) d FROM orders "
        "WHERE financial_status NOT IN ('REFUNDED','VOIDED') "
        "GROUP BY d ORDER BY d DESC LIMIT 30").fetchall()
    days = [r[0] for r in days][::-1]
    tops = [r[0] for r in db.execute(
        "SELECT title FROM order_items GROUP BY title "
        "ORDER BY SUM(quantity) DESC LIMIT 10")]

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True  # no cached results; Excel computes on open
    dash = wb.active
    dash.title = "Dashboard"

    no = len(orders) + 1   # last data row per sheet
    ni = len(items) + 1
    nv = len(inventory) + 1

    ws = _sheet(wb, "Orders",
                ["Order", "Created (UTC)", "Day (store tz)", "Month (store tz)",
                 "Status", "Total (SAR)", "Counted (SAR)"],
                orders, [10, 22, 14, 14, 12, 12, 13], {"F": NUM, "G": NUM})
    for r in range(2, no + 1):
        ws[f"G{r}"] = f'=IF(OR(E{r}="REFUNDED",E{r}="VOIDED"),0,F{r})'
        ws[f"G{r}"].font = ARIAL
        ws[f"G{r}"].number_format = NUM
    ws = _sheet(wb, "Items",
                ["Order", "Product", "Quantity", "Unit price (SAR)", "Line total (SAR)"],
                items, [10, 28, 10, 15, 15], {"D": NUM, "E": NUM})
    for r in range(2, ni + 1):
        ws[f"E{r}"] = f"=C{r}*D{r}"
        ws[f"E{r}"].font = ARIAL
        ws[f"E{r}"].number_format = NUM
    _sheet(wb, "Inventory",
           ["Product", "Status", "Variant", "SKU", "Price (SAR)", "Inventory"],
           inventory, [16, 10, 24, 10, 12, 10], {"E": NUM})

    today = datetime.date.today().isoformat()
    dash["A1"] = "Store dashboard"
    dash["A1"].font = TITLE
    dash["A2"] = (f"Synced from Shopify on {today}. Orders cover the API window "
                  "(last 60 days). Days/months in store timezone (UTC+3). "
                  "Revenue excludes refunded and voided orders.")
    dash["A4"] = "Month (blue = edit me):"
    dash["B4"] = months[-1] if months else ""
    dash["B4"].font = INPUT

    kpis = [
        ("Revenue (SAR)", f"=SUMIFS(Orders!$G$2:$G${no},Orders!$D$2:$D${no},$B$4)"),
        ("Orders", f'=COUNTIFS(Orders!$D$2:$D${no},$B$4,'
                   f'Orders!$E$2:$E${no},"<>REFUNDED",Orders!$E$2:$E${no},"<>VOIDED")'),
        ("Out-of-stock variants",
         f'=COUNTIFS(Inventory!$F$2:$F${nv},"<=0",Inventory!$B$2:$B${nv},"ACTIVE")'),
        ("Best seller", "=$A$27"),
    ]
    for i, (label, formula) in enumerate(kpis):
        col = get_column_letter(1 + i * 2)  # A C E G
        dash[f"{col}6"] = label
        dash[f"{col}6"].font = ARIAL
        dash[f"{col}7"] = formula
        dash[f"{col}7"].font = Font(name="Arial", size=14, bold=True)
        dash[f"{col}7"].number_format = NUM

    dash["A10"] = "Revenue by month"
    dash["A10"].font = BOLD
    for c, h in zip("ABC", ["Month", "Orders", "Revenue (SAR)"]):
        dash[f"{c}11"] = h
        dash[f"{c}11"].font = BOLD
    for i, m in enumerate(months):
        r = 12 + i
        dash[f"A{r}"] = m
        dash[f"B{r}"] = (f'=COUNTIFS(Orders!$D$2:$D${no},$A{r},'
                         f'Orders!$E$2:$E${no},"<>REFUNDED",Orders!$E$2:$E${no},"<>VOIDED")')
        dash[f"C{r}"] = f"=SUMIFS(Orders!$G$2:$G${no},Orders!$D$2:$D${no},$A{r})"
        dash[f"C{r}"].number_format = NUM
    m_end = 11 + len(months)

    dash["A26"] = "Top products by units"
    dash["A26"].font = BOLD
    # header row deliberately at 26 so first product sits at A27 (Best seller KPI)
    for i, t in enumerate(tops):
        r = 27 + i
        dash[f"A{r}"] = t
        dash[f"B{r}"] = f"=SUMIFS(Items!$C$2:$C${ni},Items!$B$2:$B${ni},$A{r})"
        dash[f"C{r}"] = f"=SUMIFS(Items!$E$2:$E${ni},Items!$B$2:$B${ni},$A{r})"
        dash[f"C{r}"].number_format = NUM
    t_end = 26 + len(tops)

    dash["N10"] = "Revenue by day (chart data)"
    dash["N10"].font = BOLD
    dash["N11"] = "Day"
    dash["O11"] = "Revenue (SAR)"
    for i, d in enumerate(days):
        r = 12 + i
        dash[f"N{r}"] = d
        dash[f"O{r}"] = f"=SUMIFS(Orders!$G$2:$G${no},Orders!$C$2:$C${no},$N{r})"
        dash[f"O{r}"].number_format = NUM
    d_end = 11 + len(days)

    for row in dash.iter_rows():
        for cell in row:
            if cell.font == Font():
                cell.font = ARIAL
    dash.column_dimensions["A"].width = 24
    for c in "BCDEFG":
        dash.column_dimensions[c].width = 13
    dash.column_dimensions["N"].width = 12
    dash.column_dimensions["O"].width = 13

    line = LineChart()
    line.title = "Revenue by day (SAR)"
    line.y_axis.numFmt = NUM
    line.legend = None
    line.height, line.width = 7, 16
    line.add_data(Reference(dash, min_col=15, min_row=11, max_row=d_end),
                  titles_from_data=True)
    line.set_categories(Reference(dash, min_col=14, min_row=12, max_row=d_end))
    dash.add_chart(line, "E10")

    col = BarChart()
    col.type = "col"
    col.title = "Revenue by month (SAR)"
    col.y_axis.numFmt = NUM
    col.legend = None
    col.height, col.width = 7, 16
    col.add_data(Reference(dash, min_col=3, min_row=11, max_row=m_end),
                 titles_from_data=True)
    col.set_categories(Reference(dash, min_col=1, min_row=12, max_row=m_end))
    dash.add_chart(col, "E25")

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top products by units sold"
    bar.legend = None
    bar.height, bar.width = 7, 16
    bar.add_data(Reference(dash, min_col=2, min_row=26, max_row=t_end),
                 titles_from_data=True)
    bar.set_categories(Reference(dash, min_col=1, min_row=27, max_row=t_end))
    dash.add_chart(bar, "E40")

    path = os.path.join(os.path.dirname(__file__), "shopify.xlsx")
    wb.save(path)
    # cross-check: expected values Excel should show once it recalculates
    cur = months[-1] if months else None
    exp = db.execute(
        "SELECT COUNT(*), COALESCE(SUM(total),0) FROM orders "
        "WHERE substr(datetime(created_at,'+3 hours'),1,7)=? "
        "AND financial_status NOT IN ('REFUNDED','VOIDED')", (cur,)).fetchone()
    print(f"wrote {path}")
    print(f"expected on open — {cur}: {exp[0]} orders, {exp[1]:,.0f} SAR revenue")
    return path
